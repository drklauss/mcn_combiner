import openpyxl
import datetime
import logging
import os


class Combiner:
    STATISTICS_PRICE_SEARCH_PHRASE = 'statistics'
    STATISTICS_COLUMNS_COUNT = 7
    CALLS_PRICE_SEARCH_PHRASE = 'calls'
    CALLS_COLUMNS_COUNT = 8
    RESULT_FILE_NAME = 'result.xlsx'
    _statistics_price_name = _calls_price_name = None
    _statistics = []  # Собеседник, Длительность разговора, Время звонка, Стоимость c НДС
    _calls = []  # Абонент, Собеседник,  Время звонка, Ожидание звонка
    _result_sheet = []

    @staticmethod
    def init_logger() -> None:
        """Инициализация логгера"""
        log_format = u'%(asctime)s %(levelname)s %(filename)s:%(lineno)d %(message)s'
        logging.basicConfig(filename='price_combiner.log', format=log_format, level=logging.DEBUG)

    @staticmethod
    def generate_row(stat_val: [], calls_val: []) -> tuple:
        """Создает и возвращает строку"""

        return calls_val[0], str(stat_val[0])[1:], stat_val[1], stat_val[3]

    @staticmethod
    def log_rows_info(total: int, unread: []):
        logging.warning('Rows total: {}, rows read: {}'.format(total, total - len(unread)))
        logging.warning('Unread rows: {}'.format(unread))

    def init_files(self) -> None:
        """Поиск файлов прайсов"""
        logging.info('Combiner is starting...')
        files = os.listdir('.')
        xlsx_files = tuple(filter(lambda x: '.xlsx' in x, files))
        if 2 > len(xlsx_files) or len(xlsx_files) > 2:
            raise IndexError
        for file in xlsx_files:
            if self.STATISTICS_PRICE_SEARCH_PHRASE in file:
                self._statistics_price_name = file
            if self.CALLS_PRICE_SEARCH_PHRASE in file:
                self._calls_price_name = file
        if self._statistics_price_name is None or self._calls_price_name is None:
            raise FileNotFoundError
        logging.info('Found prices: {}, {}'.format(self._statistics_price_name, self._calls_price_name))

    def _read_statistic(self):
        """Вычитывает файл статистики"""
        statistics_sheet = openpyxl.load_workbook(filename=self._statistics_price_name).active
        rows_total = 0
        unread_rows = []
        logging.info('Start reading statistics file')
        for i in statistics_sheet.rows:
            rows_total += 1
            row = tuple(filter(lambda x: x is not None, map(lambda x: x.value, i)))
            if len(row) == self.STATISTICS_COLUMNS_COUNT:
                try:
                    call_time = datetime.datetime.strptime(row[0], '%d-%m-%Y %H:%M:%S').timestamp()
                    call_duration = datetime.datetime.strptime(row[4], '%H:%M:%S').time()
                    amount_with_nds = float(str.replace(row[6], ',', '.', -1)) * 1.18
                    self._statistics.append(tuple((int(row[3]), call_duration, int(call_time), amount_with_nds)))
                except ValueError:
                    unread_rows.append(row)
                    continue

        if len(unread_rows) > 0:
            self.log_rows_info(rows_total, unread_rows)

    def _read_calls(self):
        """Вычитывает файл звонков"""
        calls_sheet = openpyxl.load_workbook(filename=self._calls_price_name).active
        rows_total = 0
        unread_rows = []
        logging.info('Start reading calls file')
        for i in calls_sheet.rows:
            rows_total += 1
            row = tuple(filter(lambda x: x is not None, map(lambda x: x.value, i)))
            if len(tuple(row)) == self.CALLS_COLUMNS_COUNT:
                try:
                    self._calls.append(tuple((int(row[1]), int(row[3]), int(row[6].timestamp()), row[5].second)))
                except ValueError:
                    unread_rows.append(row)
                    continue
        if len(unread_rows) > 0:
            self.log_rows_info(rows_total, unread_rows)

    def _reports_comparator(self):
        """Сравнивает отчеты и сохраняет результат в файл"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'result'
        not_compare = []
        for stat in self._statistics:
            success = False
            for call in self._calls:
                row = self.generate_row(stat, call)
                # Сверяем со временем ожидания + 2с, т.к. одна для окрулгения мкс, вторая для вхождения в этот интервал
                if stat[0] == call[1] and stat[2] in range(call[2], call[2] + call[3] + 2):
                    ws.append(row)
                    success = True
                    self._calls.remove(call)
                    break
            if not success:
                not_compare.append(stat)
        if len(not_compare) > 0:
            err_sum = 0.0
            for v in not_compare:
                err_sum += v[3]
            logging.warning("Not compared rows({} that costs {:.2f}): {}".format(len(not_compare), err_sum, not_compare))
        logging.info("Lost calls ({}): {}".format(len(self._calls), self._calls))
        wb.save(self.RESULT_FILE_NAME)

    def run(self):
        """Запуск """
        self.init_logger()
        self.init_files()
        self._read_statistic()
        self._read_calls()
        self._reports_comparator()
